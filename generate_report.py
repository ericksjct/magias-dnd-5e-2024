import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

FILES = {
    '2014-EN':   'js/spells-5E-2014-EN.json',
    '2014-PTBR': 'js/spells-5E-2014-PTBR.json',
    '2024-EN':   'js/spells-5E-2024-EN.json',
    '2024-PTBR': 'js/spells-5E-2024-PTBR.json',
}

# Load all spells
data = {}
for key, path in FILES.items():
    with open(path, encoding='utf-8') as f:
        data[key] = {s['name']: s for s in json.load(f)['spells']}

all_names = sorted(set(n for d in data.values() for n in d))
versions  = list(FILES.keys())

# Classify names
shared   = [n for n in all_names if sum(n in data[v] for v in versions) > 1]
orphans  = [n for n in all_names if sum(n in data[v] for v in versions) == 1]

# ── Helpers ───────────────────────────────────────────────────────────────
HDR_FILL   = PatternFill('solid', fgColor='1F3864')
HDR_FONT   = Font(bold=True, color='FFFFFF', size=11)
EVEN_FILL  = PatternFill('solid', fgColor='EEF2FF')
ODD_FILL   = PatternFill('solid', fgColor='FFFFFF')
RED_FILL   = PatternFill('solid', fgColor='FFE0E0')
GREEN_FILL = PatternFill('solid', fgColor='E0FFE8')
GRAY_FONT  = Font(color='999999', size=10)
THIN       = Side(style='thin', color='CCCCCC')
BORDER     = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

def header(ws, row, cols):
    for col, val in enumerate(cols, 1):
        c = ws.cell(row, col, val)
        c.font = HDR_FONT
        c.fill = HDR_FILL
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border = BORDER

def cell(ws, row, col, val, fill=None, font=None, bold=False, wrap=False):
    c = ws.cell(row, col, val)
    c.border = BORDER
    c.alignment = Alignment(vertical='top', wrap_text=wrap)
    if fill:  c.fill = fill
    if font:  c.font = font
    elif bold: c.font = Font(bold=True)
    return c

# ── Sheet 1: Estatísticas ─────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws_stats = wb.active
ws_stats.title = 'Estatísticas'
ws_stats.column_dimensions['A'].width = 28
ws_stats.column_dimensions['B'].width = 14

header(ws_stats, 1, ['Métrica', 'Valor'])
ws_stats.row_dimensions[1].height = 22

stats_rows = [
    ('Total de nomes únicos', len(all_names)),
    ('Nomes em mais de 1 versão', len(shared)),
    ('Nomes exclusivos (1 versão)', len(orphans)),
    (None, None),
]
for v in versions:
    stats_rows.append((f'Total em {v}', len(data[v])))
stats_rows.append((None, None))
for v in versions:
    orphan_count = sum(1 for n in orphans if n in data[v])
    stats_rows.append((f'Exclusivos de {v}', orphan_count))

for i, (label, val) in enumerate(stats_rows, 2):
    fill = EVEN_FILL if i % 2 == 0 else ODD_FILL
    if label is None:
        ws_stats.row_dimensions[i].height = 8
        continue
    cell(ws_stats, i, 1, label, fill=fill)
    c = ws_stats.cell(i, 2, val)
    c.border = BORDER
    c.alignment = Alignment(horizontal='center', vertical='top')
    c.fill = fill

# ── Sheet 2: Comparativo ──────────────────────────────────────────────────
ws_cmp = wb.create_sheet('Comparativo')

# Columns: name | present_flags | displayName per version
COL_NAME   = 1
COL_FLAGS  = 2
COL_V      = {v: 3 + i for i, v in enumerate(versions)}  # 3,4,5,6

ws_cmp.column_dimensions[get_column_letter(COL_NAME)].width  = 30
ws_cmp.column_dimensions[get_column_letter(COL_FLAGS)].width = 18
for v, col in COL_V.items():
    ws_cmp.column_dimensions[get_column_letter(col)].width = 28

hdr_cols = ['name (slug)', 'Presença'] + [f'displayName\n{v}' for v in versions]
header(ws_cmp, 1, hdr_cols)
ws_cmp.row_dimensions[1].height = 30

for i, name in enumerate(all_names, 2):
    present = [v for v in versions if name in data[v]]
    is_orphan = len(present) == 1
    fill = RED_FILL if is_orphan else (EVEN_FILL if i % 2 == 0 else ODD_FILL)

    cell(ws_cmp, i, COL_NAME, name, fill=fill, bold=True)
    cell(ws_cmp, i, COL_FLAGS, ', '.join(present), fill=fill)
    for v in versions:
        spell = data[v].get(name)
        dn = spell.get('displayName', spell.get('name', '—')) if spell else '—'
        fnt = GRAY_FONT if not spell else None
        cell(ws_cmp, i, COL_V[v], dn, fill=fill, font=fnt)

# Freeze header row
ws_cmp.freeze_panes = 'A2'
ws_stats.freeze_panes = 'A2'

# Auto-filter on comparativo
ws_cmp.auto_filter.ref = f'A1:{get_column_letter(6)}{len(all_names)+1}'

wb.save('spells_report.xlsx')
print(f'Gerado: spells_report.xlsx')
print(f'  {len(all_names)} nomes únicos | {len(shared)} compartilhados | {len(orphans)} exclusivos')
